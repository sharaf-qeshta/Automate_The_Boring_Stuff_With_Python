"""
Spreadsheet to Text Files
Write a program that performs the tasks of the previous program in reverse
order: the program should open a spreadsheet and write the cells of column
A into one text file, the cells of column B into another text file, and so on.

@author Sharaf Qeshta
"""

import openpyxl

pattern = "text"

wb = openpyxl.load_workbook("project_13_04.xlsx")
sheet = wb.active
data = []
for row in range(1, sheet.max_row + 1):
    current_row_data = []
    for col in range(1, sheet.max_column + 1):
        current_row_data.append(sheet.cell(row=row, column=col).value)
    data.append(current_row_data)

for i in range(sheet.max_column):
    for row in range(1, sheet.max_row + 1):
        try:
            file = open(f"{pattern}{i+1}.txt", "a")
            cell = sheet.cell(row=row, column=i+1)
            if cell.value is not None:
                file.write(cell.value)
            file.close()
        except Exception as error:
            print(error)
