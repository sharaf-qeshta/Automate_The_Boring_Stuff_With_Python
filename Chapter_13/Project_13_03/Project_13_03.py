"""
Spreadsheet Cell Inverter
Write a program to invert the row and column of the cells in the spreadsheet.
 For example, the value at row 5, column 3 will be at row 3, column 5
(and vice versa). This should be done for all cells in the spreadsheet. For
example, the “before” and “after” spreadsheets would look something like
Figure 13-13.
You can write this program by using nested for loops to read the
spreadsheet’s data into a list of lists data structure. This data structure
could have sheetData[x][y] for the cell at column x and row y. Then, when
writing out the new spreadsheet, use sheetData[y][x] for the cell at column
x and row y.

@author Sharaf Qeshta
"""

import openpyxl

wb = openpyxl.load_workbook("sample.xlsx")
sheet = wb.active
data = []

for row in range(1, sheet.max_row + 1):
    current_row_data = []
    for col in range(1, sheet.max_column + 1):
        current_row_data.append(sheet.cell(row=row, column=col).value)
    data.append(current_row_data)

for i in range(len(data)):
    for j in range(len(data[i])):
        cell = sheet.cell(row=i+1, column=j+1)
        cell.value = data[j][i]

wb.save("sample.xlsx")