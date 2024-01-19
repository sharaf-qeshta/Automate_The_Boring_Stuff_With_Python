"""
Multiplication Table Maker
Create a program multiplicationTable.py that takes a number N from the command
 line and creates an N×N multiplication table in an Excel spreadsheet.
For example, when the program is run like this:
py multiplicationTable.py 6
. . . it should create a spreadsheet that looks like Figure 13-11.

Row 1 and column A should be used for labels and should be in bold.

@author Sharaf Qeshta
"""

import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb.active

n = 6

for i in range(n):
    cell = sheet.cell(row=1, column=i + 2)
    cell.value = i + 1
    cell.font = Font(bold=True)
    cell = sheet.cell(row=i + 2, column=1)
    cell.value = i + 1
    cell.font = Font(bold=True)
    for j in range(n):
        cell = sheet.cell(row=i + 2, column=j + 2)
        cell.value = (i + 1) * (j + 1)

wb.save("sample.xlsx")
