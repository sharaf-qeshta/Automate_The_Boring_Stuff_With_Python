"""
Excel-to-CSV Converter
Excel can save a spreadsheet to a CSV file with a few mouse clicks, but if
you had to convert hundreds of Excel files to CSVs, it would take hours of
clicking. Using the openpyxl module from Chapter 12, write a program that
reads all the Excel files in the current working directory and outputs them
as CSV files.
A single Excel file might contain multiple sheets; you’ll have to
create one CSV file per sheet. The filenames of the CSV files should be
<excel filename>_<sheet title>.csv, where <excel filename> is the filename of
the Excel file without the file extension (for example, 'spam_data', not
'spam_data.xlsx') and <sheet title> is the string from the Worksheet object’s
title variable.
This program will involve many nested for loops. The skeleton of the
program will look something like this:
for excelFile in os.listdir('.'):
 # Skip non-xlsx files, load the workbook object.
 for sheetName in wb.get_sheet_names():
 # Loop through every sheet in the workbook.
 sheet = wb.get_sheet_by_name(sheetName)
 # Create the CSV filename from the Excel filename and sheet title.
 # Create the csv.writer object for this CSV file.
 # Loop through every row in the sheet.
 for rowNum in range(1, sheet.max_row + 1):
 rowData = [] # append each cell to this list
 # Loop through each cell in the row.
 for colNum in range(1, sheet.max_column + 1):
 # Append each cell's data to rowData.

 # Write the rowData list to the CSV file.
 csvFile.close()

 Download the ZIP file excelSpreadsheets.zip from https://nostarch.com
/automatestuff2/ and unzip the spreadsheets into the same directory as
your program. You can use these as the files to test the program on.

@author Sharaf Qeshta
"""

import os
import openpyxl
import csv

for excel_file in os.listdir('.'):
    # Skip non-xlsx files
    print(excel_file)
    if not excel_file.endswith(".xlsx"):
        continue

    # load the workbook object.
    wb = openpyxl.load_workbook(excel_file)

    for sheet_name in wb.sheetnames:  # get_sheet_names is deprecated
        # Loop through every sheet in the workbook.
        sheet = wb[sheet_name]  # get_sheet_by_name is deprecated

        # Create the CSV filename from the Excel filename and sheet title.
        csv_file = open(f"{excel_file.replace('.xlsx', '')}_{sheet_name}.csv", 'w', newline='')

        # Create the csv.writer object for this CSV file.
        csv_writer = csv.writer(csv_file)

        # Loop through every row in the sheet.
        for row_number in range(1, sheet.max_row + 1):
            row_data = []  # append each cell to this list
            # Loop through each cell in the row.
            for column_number in range(1, sheet.max_column + 1):
                # Append each cell's data to rowData.
                row_data.append(sheet.cell(row=row_number, column=column_number).value)
            # Write the rowData list to the CSV file.
            csv_writer.writerow(row_data)

        csv_file.close()
